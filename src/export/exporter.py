"""
Data Exporter for ExpiryTrack with OpenAlgo Symbol Format
"""

import io
import json
import logging
import zipfile
from collections.abc import Generator
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..database.manager import DatabaseManager

logger = logging.getLogger(__name__)

# Threshold for switching to streaming CSV export
STREAMING_ROW_THRESHOLD = 100_000


class DataExporter:
    """Export collected data in various formats with OpenAlgo symbol support"""

    def __init__(self, db_manager: DatabaseManager | None = None):
        self.db_manager = db_manager or DatabaseManager()
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)

    # ── Single JOIN helper (#13) ──────────────────────────────
    def _fetch_all_data(
        self,
        instruments: list[str],
        expiries: dict[str, list[str]],
        contract_types: list[str] | None = None,
    ) -> pd.DataFrame:
        """Fetch all data in a single JOIN query instead of N+1 per-contract queries.

        Args:
            contract_types: Optional list of contract types to include, e.g. ["CE", "PE", "FUT"].
                            When provided, only contracts matching these types are returned.
        """
        pairs = []
        for instrument in instruments:
            for exp in expiries.get(instrument, []):
                pairs.append((instrument, exp))

        if not pairs:
            return pd.DataFrame()

        with self.db_manager.get_connection() as conn:
            where_parts = []
            params: list[str] = []
            for inst, exp in pairs:
                where_parts.append("(c.instrument_key = ? AND c.expiry_date = ?)")
                params.append(inst)
                params.append(exp)

            where_clause = " OR ".join(where_parts)

            # Contract type filter
            type_clause = ""
            if contract_types:
                # Normalize: "FUT" matches anything that is NOT CE/PE
                normalized = [t.upper() for t in contract_types]
                has_fut = "FUT" in normalized
                option_types = [t for t in normalized if t in ("CE", "PE")]

                conditions = []
                if option_types:
                    placeholders = ",".join(["?"] * len(option_types))
                    conditions.append(f"c.contract_type IN ({placeholders})")
                    params.extend(option_types)
                if has_fut:
                    conditions.append("c.contract_type NOT IN ('CE', 'PE')")

                if conditions:
                    type_clause = f" AND ({' OR '.join(conditions)})"

            df = conn.execute(
                f"""
                SELECT
                    c.openalgo_symbol,
                    c.instrument_key,
                    c.expiry_date,
                    c.strike_price,
                    c.contract_type,
                    c.trading_symbol,
                    h.timestamp,
                    h.open,
                    h.high,
                    h.low,
                    h.close,
                    h.volume,
                    h.oi
                FROM historical_data h
                JOIN contracts c ON h.expired_instrument_key = c.expired_instrument_key
                WHERE ({where_clause}){type_clause}
                ORDER BY c.instrument_key, c.expiry_date, c.strike_price, h.timestamp
            """,
                params,
            ).fetchdf()

            return df

    def _add_datetime_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add date and time columns derived from timestamp."""
        if df.empty:
            return df
        ts = pd.to_datetime(df["timestamp"])
        df = df.copy()
        df["date"] = ts.dt.strftime("%Y-%m-%d")
        df["time"] = ts.dt.strftime("%H:%M:%S")
        return df

    def _apply_time_range_df(self, df: pd.DataFrame, time_range: str) -> pd.DataFrame:
        """Apply time range filter on DataFrame."""
        if df.empty or time_range == "all" or not time_range:
            return df

        days_map = {"1d": 1, "7d": 7, "30d": 30, "90d": 90}
        days = days_map.get(time_range)
        if not days:
            return df

        ts = pd.to_datetime(df["timestamp"])
        # Per-row: filter based on each row's expiry_date
        expiry_dt = pd.to_datetime(df["expiry_date"])
        cutoff = expiry_dt - pd.Timedelta(days=days)
        mask = ts >= cutoff
        return df[mask]

    def _build_filename(self, instruments: list[str], suffix: str, options: dict) -> str:
        import re as _re
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if instruments:
            raw_names = [
                inst.split("|")[1].replace(" ", "_") if "|" in inst else inst.replace(" ", "_")
                for inst in instruments
            ]
            # Strip any characters that could form path traversal sequences
            safe_names = [_re.sub(r"[^A-Za-z0-9_\-]", "", n) for n in raw_names]
            instrument_label = "_".join(safe_names[:3])
            if len(instruments) > 3:
                instrument_label += f"_and_{len(instruments) - 3}_more"
        else:
            instrument_label = "NoInstruments"
        base = f"ExpiryTrack_{instrument_label}_{timestamp}"
        if options.get("include_openalgo", True):
            base = f"OpenAlgo_{base}"
        return f"{base}.{suffix}"

    def _format_df_for_csv(self, df: pd.DataFrame, options: dict) -> pd.DataFrame:
        """Format DataFrame columns for CSV/ZIP output."""
        if df.empty:
            return df

        out = pd.DataFrame()

        if options.get("include_openalgo", True) and "openalgo_symbol" in df.columns:
            out["openalgo_symbol"] = df["openalgo_symbol"]

        out["date"] = df["date"]
        out["time"] = df["time"]

        if options.get("include_metadata", True):
            out["instrument"] = df["instrument_key"].apply(
                lambda x: x.split("|")[1].replace(" ", "_") if "|" in str(x) else x
            )
            out["expiry"] = df["expiry_date"].astype(str)
            out["strike"] = df["strike_price"]
            out["option_type"] = df["contract_type"]
            out["trading_symbol"] = df["trading_symbol"]

        out["timestamp"] = df["timestamp"]
        out["open"] = df["open"]
        out["high"] = df["high"]
        out["low"] = df["low"]
        out["close"] = df["close"]
        out["volume"] = df["volume"]
        out["oi"] = df["oi"].fillna(0).astype(int)

        return out

    def get_openalgo_formatted_symbol(self, contract: dict) -> str:
        """Convert contract to OpenAlgo symbol format"""
        try:
            symbol = contract.get("trading_symbol", "")
            if "NIFTY" in symbol:
                if "BANKNIFTY" in symbol:
                    base = "BANKNIFTY"
                elif "FINNIFTY" in symbol:
                    base = "FINNIFTY"
                else:
                    base = "NIFTY"
            else:
                base = symbol.split(" ")[0]

            expiry = contract.get("expiry_date", "")
            if expiry:
                expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
                expiry_str = expiry_date.strftime("%d%b%y").upper()
            else:
                expiry_str = ""

            if " CE " in symbol or symbol.endswith(" CE"):
                option_type = "C"
                strike = symbol.split(" CE")[0].split(" ")[-1]
            elif " PE " in symbol or symbol.endswith(" PE"):
                option_type = "P"
                strike = symbol.split(" PE")[0].split(" ")[-1]
            else:
                option_type = "F"
                strike = ""

            if option_type == "F":
                return f"{base}{expiry_str}FUT"
            else:
                return f"{base}{expiry_str}{option_type}{strike}"
        except Exception as e:
            logger.error(f"Error formatting OpenAlgo symbol: {e}")
            return contract.get("trading_symbol", "UNKNOWN")

    # ── CSV Export (rewritten with single query #13) ──────────
    def export_to_csv(self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str) -> str:
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        # Apply time range filter
        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        df = self._add_datetime_columns(df)
        df = self._format_df_for_csv(df, options)

        filename = self._build_filename(instruments, "csv", options)
        filepath = self.export_dir / filename

        if df.empty:
            headers = ["openalgo_symbol"] if options.get("include_openalgo") else []
            headers.extend(["date", "time", "timestamp", "open", "high", "low", "close", "volume", "oi"])
            pd.DataFrame(columns=headers).to_csv(filepath, index=False)
            logger.warning(f"No data found for export. Created empty file: {filepath}")
        else:
            df.to_csv(filepath, index=False)
            logger.info(f"Exported {len(df)} rows to {filepath}")

        return str(filepath)

    # ── Streaming CSV Export (for large datasets) ──────────
    def export_csv_streaming(
        self,
        instruments: list[str],
        expiries: dict[str, list[str]],
        options: dict,
    ) -> Generator[str, None, None]:
        """Generator that yields CSV data in chunks for streaming responses.

        Used when the dataset exceeds STREAMING_ROW_THRESHOLD rows to avoid
        loading the entire result set into memory at once.
        """
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        df = self._add_datetime_columns(df)
        df = self._format_df_for_csv(df, options)

        if df.empty:
            headers = ["openalgo_symbol"] if options.get("include_openalgo") else []
            headers.extend(["date", "time", "timestamp", "open", "high", "low", "close", "volume", "oi"])
            yield ",".join(headers) + "\n"
            return

        # Yield header row
        yield ",".join(df.columns) + "\n"

        # Yield data in chunks of 10,000 rows
        chunk_size = 10_000
        for start in range(0, len(df), chunk_size):
            chunk = df.iloc[start : start + chunk_size]
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False)
            yield buf.getvalue()

        logger.info(f"Streamed {len(df)} rows as CSV")

    def get_csv_row_count(
        self,
        instruments: list[str],
        expiries: dict[str, list[str]],
        options: dict,
    ) -> int:
        """Estimate the row count for a CSV export without fully materializing the data."""
        pairs = []
        for instrument in instruments:
            for exp in expiries.get(instrument, []):
                pairs.append((instrument, exp))

        if not pairs:
            return 0

        with self.db_manager.get_read_connection() as conn:
            where_parts = []
            params: list[str] = []
            for inst, exp in pairs:
                where_parts.append("(c.instrument_key = ? AND c.expiry_date = ?)")
                params.append(inst)
                params.append(exp)

            where_clause = " OR ".join(where_parts)

            row = conn.execute(
                f"""
                SELECT COUNT(*)
                FROM historical_data h
                JOIN contracts c ON h.expired_instrument_key = c.expired_instrument_key
                WHERE ({where_clause})
                """,
                params,
            ).fetchone()
            return int(row[0])

    # ── JSON Export (rewritten with single query #13) ─────────
    def export_to_json(
        self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str
    ) -> str:
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        df = self._add_datetime_columns(df)

        export_data = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "instruments": instruments,
                "format": "OpenAlgo" if options.get("include_openalgo") else "Standard",
                "version": "1.0",
            },
            "data": {},
        }

        if not df.empty:
            for instrument in instruments:
                instrument_name = instrument.split("|")[1].replace(" ", "_")
                export_data["data"][instrument_name] = {}

                inst_df = df[df["instrument_key"] == instrument]
                for expiry_date in sorted(inst_df["expiry_date"].astype(str).unique()):
                    exp_df = inst_df[inst_df["expiry_date"].astype(str) == expiry_date]
                    contracts_list = []

                    for (ts, ct, sp), group in exp_df.groupby(["trading_symbol", "contract_type", "strike_price"]):
                        contract_data = {
                            "openalgo_symbol": group["openalgo_symbol"].iloc[0]
                            if "openalgo_symbol" in group.columns
                            else "",
                            "trading_symbol": ts,
                            "strike": sp,
                            "option_type": ct,
                            "historical_data": [],
                        }
                        for _, row in group.iterrows():
                            contract_data["historical_data"].append(
                                {
                                    "date": row["date"],
                                    "time": row["time"],
                                    "timestamp": str(row["timestamp"]),
                                    "open": float(row["open"]),
                                    "high": float(row["high"]),
                                    "low": float(row["low"]),
                                    "close": float(row["close"]),
                                    "volume": int(row["volume"]),
                                    "oi": int(row["oi"]) if pd.notna(row["oi"]) else 0,
                                }
                            )
                        contracts_list.append(contract_data)

                    export_data["data"][instrument_name][expiry_date] = contracts_list

        filename = self._build_filename(instruments, "json", options)
        filepath = self.export_dir / filename

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        logger.info(f"Exported data to {filepath}")
        return str(filepath)

    # ── ZIP Export (rewritten with single query #13) ──────────
    def export_to_zip(self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str) -> str:
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        df = self._add_datetime_columns(df)

        filename = self._build_filename(instruments, "zip", options)
        zip_filepath = self.export_dir / filename

        with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
            if df.empty:
                zipf.writestr("empty.txt", "No data found for the selected parameters.")
            else:
                for instrument in instruments:
                    instrument_name = instrument.split("|")[1].replace(" ", "_")
                    inst_df = df[df["instrument_key"] == instrument]

                    for expiry_date in sorted(inst_df["expiry_date"].astype(str).unique()):
                        exp_df = inst_df[inst_df["expiry_date"].astype(str) == expiry_date]

                        if options.get("separate_files", False):
                            for ctype, label in [("CE", "CE"), ("PE", "PE"), ("FUT", "FUT")]:
                                if ctype == "FUT":
                                    type_df = exp_df[~exp_df["contract_type"].isin(["CE", "PE"])]
                                else:
                                    type_df = exp_df[exp_df["contract_type"] == ctype]
                                if not type_df.empty:
                                    formatted = self._format_df_for_csv(type_df, options)
                                    csv_name = f"{instrument_name}_{expiry_date}_{label}.csv"
                                    zipf.writestr(csv_name, formatted.to_csv(index=False))
                        else:
                            if not exp_df.empty:
                                formatted = self._format_df_for_csv(exp_df, options)
                                csv_name = f"{instrument_name}_{expiry_date}.csv"
                                zipf.writestr(csv_name, formatted.to_csv(index=False))

        logger.info(f"Created ZIP archive: {zip_filepath}")
        return str(zip_filepath)

    # ── Parquet Export (already optimized — unchanged) ────────
    def export_to_parquet(
        self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str
    ) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if instruments:
            instrument_names = [inst.split("|")[1].replace(" ", "_") for inst in instruments]
            instrument_label = "_".join(instrument_names[:3])
            if len(instruments) > 3:
                instrument_label += f"_and_{len(instruments) - 3}_more"
        else:
            instrument_label = "NoInstruments"
        base_filename = f"ExpiryTrack_{instrument_label}_{timestamp}"
        if options.get("include_openalgo", True):
            base_filename = f"OpenAlgo_{base_filename}"

        filename = f"{base_filename}.parquet"
        filepath = self.export_dir / filename

        pairs = []
        for instrument in instruments:
            for exp in expiries.get(instrument, []):
                pairs.append((instrument, exp))

        if not pairs:
            empty_df = pd.DataFrame(
                columns=[
                    "openalgo_symbol",
                    "instrument",
                    "expiry",
                    "strike",
                    "option_type",
                    "trading_symbol",
                    "timestamp",
                    "open",
                    "high",
                    "low",
                    "close",
                    "volume",
                    "oi",
                ]
            )
            empty_df.to_parquet(str(filepath), index=False)
            logger.warning(f"No data found for export. Created empty Parquet: {filepath}")
            return str(filepath)

        with self.db_manager.get_connection() as conn:
            where_parts = []
            params = []
            for inst, exp in pairs:
                where_parts.append("(c.instrument_key = ? AND c.expiry_date = ?)")
                params.append(inst)
                params.append(exp)

            where_clause = " OR ".join(where_parts)

            query = f"""
                SELECT
                    c.openalgo_symbol,
                    c.instrument_key AS instrument,
                    c.expiry_date AS expiry,
                    c.strike_price AS strike,
                    c.contract_type AS option_type,
                    c.trading_symbol,
                    h.timestamp,
                    h.open,
                    h.high,
                    h.low,
                    h.close,
                    h.volume,
                    h.oi
                FROM historical_data h
                JOIN contracts c ON h.expired_instrument_key = c.expired_instrument_key
                WHERE {where_clause}
                ORDER BY c.instrument_key, c.expiry_date, c.strike_price, h.timestamp
            """

            safe_filepath = str(filepath).replace("'", "''")
            conn.execute(
                f"""
                COPY ({query}) TO '{safe_filepath}' (FORMAT PARQUET, COMPRESSION ZSTD)
            """,
                params,
            )

            row_count = conn.execute(f"SELECT COUNT(*) FROM read_parquet('{safe_filepath}')").fetchone()[0]
            logger.info(f"Exported {row_count:,} rows to Parquet: {filepath}")

        return str(filepath)

    # ── XLSX Export (per-expiry worksheets) ─────────────────────
    def export_to_xlsx(
        self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str
    ) -> str:
        """Export data to XLSX with per-expiry worksheets, auto-column-width, and styled headers."""
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        df = self._add_datetime_columns(df)

        filename = self._build_filename(instruments, "xlsx", options)
        filepath = self.export_dir / filename

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            if df.empty:
                # Write a single empty sheet
                empty = pd.DataFrame(columns=["No data found"])
                empty.to_excel(writer, sheet_name="Empty", index=False)
            else:
                # One sheet per expiry date
                for expiry_date in sorted(df["expiry_date"].astype(str).unique()):
                    exp_df = df[df["expiry_date"].astype(str) == expiry_date]
                    formatted = self._format_df_for_csv(exp_df, options)

                    # Sheet name: max 31 chars (Excel limit)
                    sheet_name = str(expiry_date)[:31]
                    formatted.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Style the header row and auto-width
                    ws = writer.sheets[sheet_name]
                    for col_idx, col_name in enumerate(formatted.columns, 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                        # Auto-column width: max of header length and longest data value
                        max_len = len(str(col_name))
                        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 1000), min_col=col_idx, max_col=col_idx):
                            for c in row:
                                if c.value is not None:
                                    max_len = max(max_len, len(str(c.value)))
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        logger.info(f"Exported {len(df)} rows to XLSX: {filepath}")
        return str(filepath)

    # ── Amibroker Export ──────────────────────────────────────
    def export_to_amibroker(
        self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str
    ) -> str:
        """Export data in Amibroker-compatible CSV format.

        Columns: Ticker,Date/Time,Open,High,Low,Close,Volume,OpenInt
        Ticker format: SYMBOL-STRIKE-CE/PE or SYMBOL-FUT
        Date/Time format: YYYY-MM-DD HH:MM:SS
        """
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        filename = self._build_filename(instruments, "csv", options)
        # Prefix with Amibroker_ to distinguish from regular CSV
        filename = "Amibroker_" + filename
        filepath = self.export_dir / filename

        if df.empty:
            pd.DataFrame(
                columns=["Ticker", "Date/Time", "Open", "High", "Low", "Close", "Volume", "OpenInt"]
            ).to_csv(filepath, index=False)
            logger.warning(f"No data found for Amibroker export. Created empty file: {filepath}")
            return str(filepath)

        out = pd.DataFrame()
        out["Ticker"] = df.apply(self._amibroker_ticker, axis=1)
        ts = pd.to_datetime(df["timestamp"])
        out["Date/Time"] = ts.dt.strftime("%Y-%m-%d %H:%M:%S")
        out["Open"] = df["open"]
        out["High"] = df["high"]
        out["Low"] = df["low"]
        out["Close"] = df["close"]
        out["Volume"] = df["volume"]
        out["OpenInt"] = df["oi"].fillna(0).astype(int)

        out.to_csv(filepath, index=False)
        logger.info(f"Exported {len(out)} rows to Amibroker CSV: {filepath}")
        return str(filepath)

    @staticmethod
    def _amibroker_ticker(row: pd.Series) -> str:
        """Build Amibroker ticker from contract row.

        Format: SYMBOL-STRIKE-CE/PE or SYMBOL-FUT
        """
        trading_symbol = str(row.get("trading_symbol", ""))
        contract_type = str(row.get("contract_type", ""))
        strike = row.get("strike_price", "")

        # Extract base symbol from instrument_key (e.g. "NSE_INDEX|Nifty 50" -> "NIFTY50")
        inst_key = str(row.get("instrument_key", ""))
        if "|" in inst_key:
            symbol = inst_key.split("|")[1].replace(" ", "").upper()
        elif trading_symbol:
            symbol = trading_symbol.split(" ")[0].upper()
        else:
            symbol = "UNKNOWN"

        if contract_type in ("CE", "PE"):
            strike_str = str(int(float(strike))) if strike else "0"
            return f"{symbol}-{strike_str}-{contract_type}"
        else:
            return f"{symbol}-FUT"

    # ── MetaTrader Export ─────────────────────────────────────
    def export_to_metatrader(
        self, instruments: list[str], expiries: dict[str, list[str]], options: dict, task_id: str
    ) -> str:
        """Export data in MetaTrader (MT4/MT5) compatible CSV format.

        Columns: Date,Time,Open,High,Low,Close,Volume
        Date format: YYYY.MM.DD
        Time format: HH:MM
        """
        df = self._fetch_all_data(instruments, expiries, contract_types=options.get("contract_types"))

        if options.get("time_range") != "all":
            df = self._apply_time_range_df(df, options.get("time_range"))

        filename = self._build_filename(instruments, "csv", options)
        filename = "MetaTrader_" + filename
        filepath = self.export_dir / filename

        if df.empty:
            pd.DataFrame(
                columns=["Date", "Time", "Open", "High", "Low", "Close", "Volume"]
            ).to_csv(filepath, index=False)
            logger.warning(f"No data found for MetaTrader export. Created empty file: {filepath}")
            return str(filepath)

        out = pd.DataFrame()
        ts = pd.to_datetime(df["timestamp"])
        out["Date"] = ts.dt.strftime("%Y.%m.%d")
        out["Time"] = ts.dt.strftime("%H:%M")
        out["Open"] = df["open"]
        out["High"] = df["high"]
        out["Low"] = df["low"]
        out["Close"] = df["close"]
        out["Volume"] = df["volume"]

        out.to_csv(filepath, index=False)
        logger.info(f"Exported {len(out)} rows to MetaTrader CSV: {filepath}")
        return str(filepath)

    def get_available_expiries(self, instruments: list[str]) -> dict[str, list[str]]:
        result = {}
        for instrument in instruments:
            expiries = self.db_manager.get_expiries_for_instrument(instrument)
            result[instrument] = sorted(expiries, reverse=True)
        return result

    # ── Candle Data Export ────────────────────────────────────

    def _fetch_candle_data(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
    ) -> pd.DataFrame:
        """Fetch candle data with optional instrument master join"""
        if not instrument_keys:
            return pd.DataFrame()

        with self.db_manager.get_connection() as conn:
            placeholders = ",".join(["?"] * len(instrument_keys))
            params = list(instrument_keys) + [interval]

            where_extra = ""
            if from_date:
                where_extra += " AND cd.timestamp >= ?"
                params.append(from_date)
            if to_date:
                where_extra += " AND cd.timestamp <= ?"
                params.append(to_date)

            df = conn.execute(
                f"""
                SELECT
                    cd.instrument_key,
                    COALESCE(im.trading_symbol, cd.instrument_key) as symbol,
                    COALESCE(im.name, '') as name,
                    COALESCE(im.segment, '') as segment,
                    cd.timestamp,
                    cd.open,
                    cd.high,
                    cd.low,
                    cd.close,
                    cd.volume,
                    cd.oi,
                    cd.interval
                FROM candle_data cd
                LEFT JOIN instrument_master im ON cd.instrument_key = im.instrument_key
                WHERE cd.instrument_key IN ({placeholders})
                  AND cd.interval = ?
                  {where_extra}
                ORDER BY cd.instrument_key, cd.timestamp
            """,
                params,
            ).fetchdf()

            return df

    def export_candles_to_csv(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
        task_id: str | None = None,
    ) -> str:
        """Export candle data to CSV"""
        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments")

        df = self._add_datetime_columns(df)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"candles_{interval}_{timestamp}.csv"
        filepath = self.export_dir / filename

        df.to_csv(filepath, index=False)
        logger.info(f"Exported {len(df):,} candle rows to CSV: {filepath}")
        return str(filepath)

    def export_candles_to_parquet(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
        task_id: str | None = None,
    ) -> str:
        """Export candle data to Parquet"""
        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"candles_{interval}_{timestamp}.parquet"
        filepath = self.export_dir / filename

        df.to_parquet(filepath, index=False, engine="pyarrow")
        logger.info(f"Exported {len(df):,} candle rows to Parquet: {filepath}")
        return str(filepath)

    def _export_candles_to_json(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
    ) -> str:
        """Export candle data grouped by instrument key to JSON."""
        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments")

        result: dict = {}
        for key, group in df.groupby("instrument_key"):
            g = group[["timestamp", "open", "high", "low", "close", "volume", "oi"]].copy()
            g["timestamp"] = g["timestamp"].astype(str)
            g["volume"] = g["volume"].fillna(0).astype(int)
            g["oi"] = g["oi"].fillna(0).astype(int)
            result[str(key)] = g.to_dict("records")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"candles_{interval}_{ts}.json"
        with open(filepath, "w") as f:
            json.dump(result, f, indent=2, default=str)

        logger.info(f"Exported {len(df):,} candle rows to JSON: {filepath}")
        return str(filepath)

    def _export_candles_to_xlsx(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
    ) -> str:
        """Export candle data to XLSX with one worksheet per instrument."""
        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"candles_{interval}_{ts}.xlsx"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for key, group in df.groupby("instrument_key"):
                inst_df = group.copy()
                symbol = inst_df["symbol"].iloc[0] if "symbol" in inst_df.columns else str(key)
                sheet_name = str(symbol)[:31]
                inst_df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]
                for col_idx, col_name in enumerate(inst_df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    max_len = len(str(col_name))
                    for row in ws.iter_rows(
                        min_row=2, max_row=min(ws.max_row, 1000),
                        min_col=col_idx, max_col=col_idx
                    ):
                        for c in row:
                            if c.value is not None:
                                max_len = max(max_len, len(str(c.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        logger.info(f"Exported {len(df):,} candle rows to XLSX: {filepath}")
        return str(filepath)

    def _export_candles_to_zip(
        self,
        instrument_keys: list[str],
        interval: str = "1day",
        from_date: str | None = None,
        to_date: str | None = None,
    ) -> str:
        """Export candle data to ZIP with one CSV file per instrument."""
        import re as _re

        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = self.export_dir / f"candles_{interval}_{ts}.zip"

        with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
            for key, group in df.groupby("instrument_key"):
                inst_df = group.copy()
                symbol = inst_df["symbol"].iloc[0] if "symbol" in inst_df.columns else str(key)
                safe_name = _re.sub(r"[^A-Za-z0-9_\-]", "_", str(symbol))
                csv_name = f"{safe_name}_{interval}.csv"
                zipf.writestr(csv_name, inst_df.to_csv(index=False))

        logger.info(f"Exported {len(df):,} candle rows to ZIP: {filepath}")
        return str(filepath)

    def export_candles_bulk(
        self,
        instrument_keys: list[str],
        interval: str,
        export_format: str,
        from_date: str | None = None,
        to_date: str | None = None,
    ) -> tuple[str, int]:
        """Dispatch candle export to the correct format. Returns (file_path, row_count)."""
        # Pre-fetch to validate non-empty and capture row count
        df = self._fetch_candle_data(instrument_keys, interval, from_date, to_date)
        if df.empty:
            raise ValueError("No candle data found for the selected instruments and filters")
        row_count = len(df)

        if export_format == "csv":
            path = self.export_candles_to_csv(instrument_keys, interval, from_date, to_date)
        elif export_format == "parquet":
            path = self.export_candles_to_parquet(instrument_keys, interval, from_date, to_date)
        elif export_format == "json":
            path = self._export_candles_to_json(instrument_keys, interval, from_date, to_date)
        elif export_format == "xlsx":
            path = self._export_candles_to_xlsx(instrument_keys, interval, from_date, to_date)
        elif export_format == "zip":
            path = self._export_candles_to_zip(instrument_keys, interval, from_date, to_date)
        else:
            raise ValueError(f"Unsupported candle format: {export_format}")

        return path, row_count

    def export_instrument_master(
        self,
        export_format: str,
        segment: str | None = None,
    ) -> tuple[str, int]:
        """Export instrument master catalog. Returns (file_path, row_count)."""
        with self.db_manager.get_read_connection() as conn:
            sql = "SELECT * FROM instrument_master"
            params: list[str] = []
            if segment:
                sql += " WHERE segment = ?"
                params.append(segment)
            sql += " ORDER BY segment, instrument_type, trading_symbol"
            df = conn.execute(sql, params).fetchdf()

        if df.empty:
            raise ValueError("No instrument master data found. Import instruments first.")

        suffix = f"_{segment}" if segment else "_all"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"instrument_master{suffix}_{ts}"

        if export_format == "csv":
            path = self.export_dir / f"{filename}.csv"
            df.to_csv(path, index=False)
        elif export_format == "parquet":
            path = self.export_dir / f"{filename}.parquet"
            df.to_parquet(path, index=False, engine="pyarrow")
        elif export_format == "json":
            path = self.export_dir / f"{filename}.json"
            df.to_json(path, orient="records", indent=2)
        else:
            raise ValueError(f"Unsupported format: {export_format}")

        logger.info(f"Exported {len(df):,} instrument master rows to {export_format}: {path}")
        return str(path), len(df)
